import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path

import paramiko
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
SOURCE = Path('Ma_tran_phan_quyen_KiotViet_Phu_tung_o_to.xlsx')
IMPORTER = Path('cooling-php/_vps_live/import_rbac_phase31.php')

def value(cell):
    return str(cell).strip() if cell is not None else ''

def access(value_text):
    normalized = value_text.replace('–', '—').strip()
    return {'TQ':'TQ','QL':'QL','TH':'TH','X':'X','—':'NONE','-':'NONE','':'NONE'}.get(normalized, None)

book = load_workbook(SOURCE, data_only=True, read_only=True)
roles_sheet = book['Vai trò']
matrix_sheet = book['Ma trận phân quyền']

role_header = next(row for row in roles_sheet.iter_rows(values_only=True) if value(row[0]) == 'Mã')
roles = []
for row in roles_sheet.iter_rows(values_only=True):
    code = value(row[0])
    if code in {'SA','OWN','BM','SAL','CAS','WH','PUR','ACC','CS','TECH','DEL','MKT','AUD'}:
        roles.append({'code':code,'name':value(row[1]),'default_scope':value(row[2]),'responsibility':value(row[3]),'restricted_actions':value(row[4]),'implementation_note':value(row[5])})
role_codes = [role['code'] for role in roles]
if len(roles) != 13 or len(set(role_codes)) != 13:
    raise RuntimeError(f'Expected 13 unique roles, found {len(roles)}')

header_index = None
headers = None
for index, row in enumerate(matrix_sheet.iter_rows(values_only=True), start=1):
    if value(row[0]) == 'STT' and value(row[1]) == 'Phân hệ':
        header_index = index
        headers = [value(cell) for cell in row]
        break
if header_index is None: raise RuntimeError('Permission matrix header not found')
role_columns = {code: headers.index(code) for code in role_codes}
permissions = []
matrix = []
for row in list(matrix_sheet.iter_rows(values_only=True))[header_index:]:
    raw_id = row[0]
    try:
        order = int(raw_id)
    except (TypeError, ValueError):
        continue
    module = value(row[1]); feature = value(row[2]); action_name = value(row[3])
    module_code = module.split('.', 1)[0].zfill(2) if module else '00'
    permission_code = f'P{order:03d}'
    permissions.append({'code':permission_code,'module_code':module_code,'module_name':module,'feature_name':feature,'action_name':action_name,'data_scope':value(row[4]),'sensitivity':value(row[5]),'note':value(row[6]),'sort_order':order})
    for role_code, column in role_columns.items():
        level = access(value(row[column]))
        if level is None: raise RuntimeError(f'Unknown access level at permission {order}, role {role_code}: {value(row[column])!r}')
        matrix.append({'role_code':role_code,'permission_code':permission_code,'access_level':level})
if len(permissions) != 157 or len(matrix) != 13 * 157:
    raise RuntimeError(f'Unexpected matrix size: {len(permissions)} permissions, {len(matrix)} entries')

payload = {'schema_version':'rbac-phase-3.1','source_file':SOURCE.name,'source_sha256':hashlib.sha256(SOURCE.read_bytes()).hexdigest(),'roles':roles,'permissions':permissions,'matrix':matrix}

client = paramiko.SSHClient(); client.set_missing_host_key_policy(paramiko.AutoAddPolicy()); client.connect('103.97.134.164', username='root', password='lcBFDjVF15', timeout=30)
def run(command, timeout=60):
    _, out, err = client.exec_command(command, timeout=timeout)
    output = out.read().decode('utf-8','replace'); error = err.read().decode('utf-8','replace'); code = out.channel.recv_exit_status()
    if code: raise RuntimeError(f'{command}\n{output}\n{error}')
    return output.strip()

stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
backup = f'/var/backups/coolingsystems/rbac-phase31-{stamp}'
baseline = run("sqlite3 -separator '|' /var/lib/coolingsystems/cooling.db \"SELECT (SELECT COUNT(*) FROM products),(SELECT COUNT(*) FROM users),(SELECT COUNT(*) FROM staff_roles),(SELECT COUNT(*) FROM staff_role_assignments),(SELECT COUNT(*) FROM orders);\"")
run(f"mkdir -p {backup} && sqlite3 /var/lib/coolingsystems/cooling.db '.backup {backup}/cooling.db' && cp /opt/coolingsystems/includes/auth.php {backup}/auth.php && cp /opt/coolingsystems/routes/admin.php {backup}/admin.php && cp /opt/coolingsystems/views/partials/dashboard-head.php {backup}/dashboard-head.php")

sftp = client.open_sftp()
remote_json = '/tmp/coolingsystems-rbac-phase31.json'
remote_importer = '/tmp/coolingsystems-rbac-phase31.php'
with sftp.open(remote_json,'wb') as remote: remote.write(json.dumps(payload, ensure_ascii=False).encode('utf-8'))
sftp.put(str(IMPORTER), remote_importer)
sftp.close()
result = run(f"chown www-data:www-data {remote_json} {remote_importer} && runuser -u www-data -- php {remote_importer} {remote_json}; rm -f {remote_json} {remote_importer}")
after = run("sqlite3 -separator '|' /var/lib/coolingsystems/cooling.db \"SELECT (SELECT COUNT(*) FROM products),(SELECT COUNT(*) FROM users),(SELECT COUNT(*) FROM staff_roles),(SELECT COUNT(*) FROM staff_role_assignments),(SELECT COUNT(*) FROM orders),(SELECT COUNT(*) FROM rbac_roles),(SELECT COUNT(*) FROM rbac_permissions),(SELECT COUNT(*) FROM rbac_role_permissions);\"")
print('BACKUP=' + backup)
print('BASELINE=' + baseline)
print('IMPORT=' + result)
print('AFTER=' + after)
client.close()
