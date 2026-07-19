from pathlib import Path
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
path = Path('Ma_tran_phan_quyen_KiotViet_Phu_tung_o_to.xlsx')
book = load_workbook(path, data_only=False, read_only=True)
print('SHEETS:', ', '.join(book.sheetnames))

for sheet in book.worksheets:
    print(f'\n--- {sheet.title} ({sheet.max_row} rows x {sheet.max_column} cols) ---')
    for row in sheet.iter_rows(values_only=True):
        values = [str(value).strip() if value is not None else '' for value in row]
        if any(values):
            print(' | '.join(values))
