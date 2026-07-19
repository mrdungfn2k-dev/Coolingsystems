import re
import sys
from pathlib import Path

import paramiko
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
book = load_workbook(Path('Ma_tran_phan_quyen_KiotViet_Phu_tung_o_to.xlsx'), read_only=True, data_only=True)
sheet = book['Ma trận phân quyền']
excel = {}
for row in sheet.iter_rows(values_only=True):
    first = row[0]
    if isinstance(first, int) and 1 <= first <= 999:
        excel[first] = {
            'group': str(row[1] or '').strip(),
            'feature': str(row[2] or '').strip(),
            'action': str(row[3] or '').strip(),
        }

client = paramiko.SSHClient()
client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
client.connect('103.97.134.164', username='root', password='lcBFDjVF15', timeout=30)


def run(command):
    _, out, err = client.exec_command(command, timeout=60)
    stdout = out.read().decode('utf-8', 'replace')
    stderr = err.read().decode('utf-8', 'replace')
    if out.channel.recv_exit_status():
        raise RuntimeError(command + '\n' + stdout + '\n' + stderr)
    return stdout.strip()


rows = run("sqlite3 -separator '|' /var/lib/coolingsystems/cooling.db \"SELECT permission_code,capability FROM rbac_capability_rules ORDER BY CAST(SUBSTR(permission_code,2) AS INTEGER),capability;\"")
rules = {}
for line in rows.splitlines():
    code, capability = (line.split('|', 1) + [''])[:2]
    match = re.fullmatch(r'P(\d+)', code)
    if match:
        rules.setdefault(int(match.group(1)), []).append(capability)

tables = run("sqlite3 /var/lib/coolingsystems/cooling.db \"SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;\"").splitlines()
client.close()

print('EXCEL_COUNT=' + str(len(excel)))
print('EXCEL_RANGE=' + str(min(excel)) + '-' + str(max(excel)))
print('RULE_CODE_COUNT=' + str(len(set(excel) & set(rules))))
print('RULE_MISSING=' + ','.join('P' + str(code) for code in sorted(set(excel) - set(rules))))
print('RULE_PRESENT=' + ','.join('P' + str(code) for code in sorted(set(excel) & set(rules))))
print('BUSINESS_TABLES=' + ','.join(name for name in tables if any(term in name for term in ('warranty', 'cash_', 'supplier', 'purchase_', 'debt', 'bank_reconciliation'))))
for code in range(77, 128):
    if code in excel:
        item = excel[code]
        print(f"P{code}|{item['group']}|{item['feature']}|{item['action']}|{','.join(rules.get(code, [])) or '-'}")
